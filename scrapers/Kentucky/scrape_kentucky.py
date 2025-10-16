import cloudscraper
from bs4 import BeautifulSoup
import pandas as pd
import re
import os
from datetime import datetime
import openpyxl
import logging
import subprocess
import signal
import sys

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Global variable to store caffeinate process
caffeinate_process = None

def prevent_sleep():
    """Prevent the computer from sleeping during scraping"""
    global caffeinate_process
    try:
        # Use caffeinate to prevent sleep
        caffeinate_process = subprocess.Popen(['caffeinate', '-i'])
        logging.info("Computer sleep prevention activated")
    except Exception as e:
        logging.warning(f"Could not prevent sleep: {e}")

def allow_sleep():
    """Allow the computer to sleep again"""
    global caffeinate_process
    if caffeinate_process:
        try:
            caffeinate_process.terminate()
            logging.info("Computer sleep prevention deactivated")
        except Exception as e:
            logging.warning(f"Could not deactivate sleep prevention: {e}")

def signal_handler(signum, frame):
    """Handle interrupt signals to clean up"""
    logging.info("Received interrupt signal, cleaning up...")
    allow_sleep()
    sys.exit(0)

# Register signal handlers
signal.signal(signal.SIGINT, signal_handler)
signal.signal(signal.SIGTERM, signal_handler)

def cleanup_old_files(script_dir, file_pattern, keep=1):
    """Delete old files matching the pattern, keeping the most recent 'keep' files."""
    try:
        import glob
        import os
        old_files = glob.glob(os.path.join(script_dir, file_pattern))
        # Sort files by modification time, newest last
        old_files.sort(key=lambda x: os.path.getmtime(x))
        # Keep only the most recent 'keep' files
        files_to_delete = old_files[:-keep]
        for old_file in files_to_delete:
            try:
                os.remove(old_file)
                logging.info(f"Deleted old file: {os.path.basename(old_file)}")
            except Exception as e:
                logging.warning(f"Could not delete {old_file}: {e}")
    except Exception as e:
        logging.warning(f"Error during cleanup: {e}")

def save_progress(candidates_by_year, script_dir, progress_name):
    """Save progress to prevent data loss - creates one combined progress file, keeps two most recent backups."""
    try:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # Clean up old progress files before saving new ones, keep two most recent
        cleanup_old_files(script_dir, "progress_*.xlsx", keep=2)
        
        # Combine all candidates into one progress file
        all_candidates = []
        for year, candidates in candidates_by_year.items():
            if candidates:
                all_candidates.extend(candidates)
        
        if all_candidates:
            # Create DataFrame with all candidates
            df = pd.DataFrame(all_candidates)
            
            # Create single progress file
            progress_filename = f'progress_{progress_name}_{timestamp}.xlsx'
            progress_path = os.path.join(script_dir, progress_filename)
            
            # Save to Excel
            with pd.ExcelWriter(progress_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Progress')
                
                workbook = writer.book
                worksheet = writer.sheets['Progress']
                
                # Format headers
                for cell in worksheet[1]:
                    cell.font = cell.font.copy(bold=True)
                    cell.fill = openpyxl.styles.PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = openpyxl.utils.get_column_letter(column[0].column)
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            logging.info(f"Saved progress: {progress_filename} with {len(all_candidates)} total candidates")
        
    except Exception as e:
        logging.error(f"Error saving progress: {e}")

def extract_year_from_date(date_str):
    """Extract year from date string in format MM/DD/YYYY"""
    try:
        if date_str and date_str.strip():
            # Handle different date formats
            if '/' in date_str:
                parts = date_str.split('/')
                if len(parts) == 3:
                    return int(parts[2])
            elif '-' in date_str:
                parts = date_str.split('-')
                if len(parts) == 3:
                    return int(parts[2])
        return None
    except:
        return None

def parse_candidate_row(row):
    """Parse a candidate row and extract relevant information"""
    try:
        cells = row.find_all(['td', 'th'])
        if len(cells) < 6:
            return None
        # The first cell (cells[0]) is a link/empty, so skip it
        # cells[1]: Office
        # cells[3]: Location
        # cells[4]: Date
        # cells[5]: Election
        office = cells[1].get_text(strip=True) if len(cells) > 1 else ""
        location = cells[3].get_text(strip=True) if len(cells) > 3 else ""
        election_date = cells[4].get_text(strip=True) if len(cells) > 4 else ""
        election_type = cells[5].get_text(strip=True) if len(cells) > 5 else ""
        # Extract name from the first cell (cells[0], which is a link)
        name_link = cells[0].find('a')
        if name_link:
            name_text = name_link.get_text(strip=True)
            name_parts = name_text.split()
            if len(name_parts) >= 2:
                first_name = name_parts[0]
                last_name = ' '.join(name_parts[1:])
            else:
                first_name = name_text
                last_name = ""
        else:
            first_name = ""
            last_name = ""
        return {
            'last_name': last_name,
            'first_name': first_name,
            'office_sought': office,
            'location': location,
            'election_date': election_date,
            'election_type': election_type
        }
    except Exception as e:
        logging.error(f"Error parsing row: {e}")
        return None

def scrape_kentucky_candidates():
    """Scrape Kentucky candidate data and organize by election years"""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    
    # Activate sleep prevention
    prevent_sleep()
    
    try:
        # Base URL for Kentucky candidate search
        base_url = "https://secure.kentucky.gov/kref/publicsearch/CandidateSearch"
        
        # Initialize data storage by year
        candidates_by_year = {
            2025: [],
            2024: [],
            2023: [],
            2022: [],
            2020: [],
            'other': []
        }
        
        # Create a cloudscraper session
        scraper = cloudscraper.create_scraper(
            browser={
                'browser': 'chrome',
                'platform': 'darwin',
                'desktop': True
            }
        )
        
        page_num = 1
        total_candidates = 0
        last_page_num = None
        
        while True:
            logging.info(f"Scraping page {page_num}...")
            
            # Construct URL with pagination
            if page_num == 1:
                url = f"{base_url}?FirstName=&LastName=&ElectionDate=01%2F01%2F0001&PoliticalParty=&ExemptionStatus=All&IsActiveFlag="
            else:
                # Use pageIndex parameter (0-based) for pagination
                page_index = page_num - 1  # Convert to 0-based index
                url = f"{base_url}?FirstName=&LastName=&ElectionDate=01%2F01%2F0001&PoliticalParty=&ExemptionStatus=All&IsActiveFlag=&pageIndex={page_index}&pageSize=10"
            
            # Get the page content
            response = scraper.get(url)
            response.raise_for_status()
            
            # Parse HTML
            soup = BeautifulSoup(response.text, 'html.parser')
            
            # On the first page, determine the last page number from totalRecord and pageSize in the pagination links
            if page_num == 1:
                pagination = soup.find('ul', class_='pagination')
                if pagination:
                    page_links = pagination.find_all('a', class_='page-link')
                    total_record = None
                    page_size = None
                    for link in page_links:
                        href = link.get('href', '')
                        if 'totalRecord=' in href and 'pageSize=' in href:
                            import re
                            total_record_match = re.search(r'totalRecord=(\d+)', href)
                            page_size_match = re.search(r'pageSize=(\d+)', href)
                            if total_record_match:
                                total_record = int(total_record_match.group(1))
                            if page_size_match:
                                page_size = int(page_size_match.group(1))
                    if total_record and page_size:
                        import math
                        last_page_num = math.ceil(total_record / page_size)
                        logging.info(f"Detected last page number from totalRecord/pageSize: {last_page_num}")
            
            # Find the table
            table = soup.find('table')
            if not table:
                logging.warning(f"No table found on page {page_num}")
                break
            
            # Find all rows in the table body
            rows = table.find_all('tr')
            if not rows:
                logging.warning(f"No rows found on page {page_num}")
                break

            # If only the header row is present, stop
            if len(rows) <= 1:
                logging.info(f"No candidate data found on page {page_num}. Reached the last page. Stopping.")
                break

            page_candidates = 0
            
            # Process each row (skip header row)
            for row in rows[1:]:  # Skip header row
                candidate_data = parse_candidate_row(row)
                if candidate_data:
                    # Extract year from election date
                    year = extract_year_from_date(candidate_data['election_date'])
                    
                    # Categorize by year
                    if year in candidates_by_year:
                        candidates_by_year[year].append(candidate_data)
                    else:
                        candidates_by_year['other'].append(candidate_data)
                    
                    page_candidates += 1
                    total_candidates += 1
            
            logging.info(f"Found {page_candidates} candidates on page {page_num}")
            
            # Save progress every 500 pages to prevent data loss
            if page_num % 500 == 0:
                logging.info(f"Saving progress after {page_num} pages...")
                save_progress(candidates_by_year, script_dir, f"progress_page_{page_num}")
            
            # Stop if we've reached the last page
            if last_page_num is not None and page_num >= last_page_num:
                logging.info(f"Reached the last page ({last_page_num}). Stopping.")
                break
            
            # Check if there's a next page (legacy fallback)
            next_link = soup.find('a', string='»')
            if not next_link or 'disabled' in next_link.get('class', []):
                logging.info("No more pages to scrape")
                break
            
            page_num += 1
            
            # Add a small delay to be respectful to the server
            import time
            time.sleep(1)
        
        logging.info(f"Total candidates scraped: {total_candidates}")
        
        # Clean up old final Excel files before creating new ones
        cleanup_old_files(script_dir, "kentucky_candidates_*.xlsx")
        
        # Create one combined Excel file with all data
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        all_candidates = []
        for candidates in candidates_by_year.values():
            all_candidates.extend(candidates)
        
        if all_candidates:
            df_all = pd.DataFrame(all_candidates)
            combined_filename = f'kentucky_candidates_{timestamp}.xlsx'
            raw_dir = os.path.join('data', 'raw')
            os.makedirs(raw_dir, exist_ok=True)
            combined_path = os.path.join(raw_dir, combined_filename)
            
            with pd.ExcelWriter(combined_path, engine='openpyxl') as writer:
                df_all.to_excel(writer, index=False, sheet_name='Candidates')
                
                workbook = writer.book
                worksheet = writer.sheets['Candidates']
                
                # Format headers
                for cell in worksheet[1]:
                    cell.font = cell.font.copy(bold=True)
                    cell.fill = openpyxl.styles.PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = openpyxl.utils.get_column_letter(column[0].column)
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            logging.info(f"Created combined Excel file: {combined_filename} with {len(all_candidates)} total candidates")
        
    except Exception as e:
        logging.error(f"Error: {e}")
        raise
    finally:
        # Always deactivate sleep prevention when done
        allow_sleep()

if __name__ == "__main__":
    scrape_kentucky_candidates() 